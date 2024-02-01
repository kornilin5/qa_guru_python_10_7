import zipfile
import os
import shutil
from conftest import NAME_FILES, RESOURCES_PATH, TMP_PATH
from openpyxl import load_workbook
from pypdf import PdfReader
import csv

def create_archive(file_names, archive_path, tmp_path):
    files_to_zip = [os.path.join(tmp_path, file_name) for file_name in file_names]
    zip_file_name = os.path.join(archive_path, 'achive_test.zip')

    with zipfile.ZipFile(zip_file_name, 'w') as zip_file:
        for file in files_to_zip:
            zip_file.write(file, os.path.relpath(file, tmp_path))
    print(f'Zip-архив "{zip_file_name}" создан успешно.')

create_archive(NAME_FILES, RESOURCES_PATH, TMP_PATH)


def extract_file_from_zip(zip_path, target_folder, target_file):
    zip_file_path = os.path.join(zip_path, 'achive_test.zip')
    with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
        zip_ref.extract(target_file, target_folder)


def read_csv(path_archive, folder_path, name_file):
    extract_file_from_zip(path_archive, folder_path, name_file)
    path = os.path.join(folder_path, name_file)
    target_text = 'eto test'
    with open(path, encoding='utf-8') as csvfile:
        found_rows = [row for row in csv.reader(csvfile, delimiter=';') if any(target_text in value for value in row)]
    for row in found_rows:
        print(f"Найден текст '{target_text}'")

read_csv(RESOURCES_PATH,RESOURCES_PATH, 'это_csv.csv')

def read_xlsx_file(path_archive, folder_path,  name_file):
    extract_file_from_zip(path_archive, folder_path, name_file)
    path = os.path.join(folder_path, name_file)
    open_xlsx = load_workbook(path)
    sheet = open_xlsx.active
    text = sheet.cell(row=2, column=2).value
    text_found = "Регистрация пользователя на сайте при вводе корректного пароля"
    assert text == text_found

read_xlsx_file(RESOURCES_PATH,RESOURCES_PATH, 'эксель.xlsx')


def read_pdf(path_archive, folder_path, name_file):
    extract_file_from_zip(path_archive, folder_path, name_file)
    path = os.path.join(folder_path, name_file)
    text = "Пример PDF файла"
    with open(path, 'rb') as file:
        reader = PdfReader(file)
        print(reader)
        assert any(text in page.extract_text() for page in reader.pages)


read_pdf(RESOURCES_PATH, RESOURCES_PATH, 'пдф.pdf')
